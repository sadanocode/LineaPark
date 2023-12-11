import openpyxl
import os.path
import settings
from loguru import logger as cs_logger
from sys import stderr


cs_logger.remove()
cs_logger.add(stderr, format="<white>{time:HH:mm:ss}</white> | <level>{level: <8}</level> | <white>{message}</white>")


def create_xml():
    log_file = settings.log_file
    if os.path.exists(log_file) is False:
        workbook = openpyxl.Workbook()
        workbook.save(log_file)

        worksheet = workbook.active
        worksheet.title = 'Overall'
        worksheet.cell(row=1, column=1).value = "№ кошелька"
        worksheet.cell(row=1, column=2).value = "Адрес кошелька"
        worksheet.cell(row=1, column=3).value = "Адрес биржи"
        worksheet.cell(row=1, column=4).value = "Сумма бриджа"
        worksheet.cell(row=1, column=5).value = "Нач. бал Linea"
        worksheet.cell(row=1, column=6).value = "Кон. бал Linea"
        worksheet.cell(row=1, column=7).value = "Кол-во транз в скрипте"
        worksheet.cell(row=1, column=8).value = "Кол-во транз кошелька"
        worksheet.cell(row=1, column=9).value = "Нач. баланс биржи"
        worksheet.cell(row=1, column=10).value = "Кон. баланс биржи"
        worksheet.cell(row=1, column=11).value = "Сумма депозита $ FWDX"
        worksheet.cell(row=1, column=12).value = "Сумма депозита $ ZKDX"
        worksheet.cell(row=1, column=13).value = "Время"
        workbook.save(log_file)

        worksheet = workbook.create_sheet('Bridge transactions')
        worksheet.cell(row=1, column=1).value = "№ кошелька"
        worksheet.cell(row=1, column=2).value = "Адрес"
        worksheet.cell(row=1, column=3).value = "Исх. сеть"
        worksheet.cell(row=1, column=4).value = "Вхд. сеть"
        worksheet.cell(row=1, column=5).value = f"Hash бриджа"
        worksheet.cell(row=1, column=6).value = f"Нач. баланс исх. сети"
        worksheet.cell(row=1, column=7).value = f"Кон. баланс исх. сети"
        worksheet.cell(row=1, column=8).value = "Сумма бриджа ETH"
        worksheet.cell(row=1, column=9).value = f"Нач. баланс вхд. сети"
        worksheet.cell(row=1, column=10).value = f"Кон. баланс вхд. сети"
        worksheet.cell(row=1, column=11).value = "Время"
        workbook.save(log_file)
        workbook.close()

        worksheet = workbook.create_sheet('Swap transactions')
        worksheet.cell(row=1, column=1).value = "№ кошелька"
        worksheet.cell(row=1, column=2).value = "№ транзакции"
        worksheet.cell(row=1, column=3).value = "Адрес кошелька"
        worksheet.cell(row=1, column=4).value = "Свап"
        worksheet.cell(row=1, column=5).value = "Сумма свапа ETH"
        worksheet.cell(row=1, column=6).value = "Сумма свапа токена"
        worksheet.cell(row=1, column=7).value = "Hash свапа"
        worksheet.cell(row=1, column=8).value = "Нач. баланс ETH"
        worksheet.cell(row=1, column=9).value = "Кон. баланс ETH"
        worksheet.cell(row=1, column=10).value = "Нач. баланс токена"
        worksheet.cell(row=1, column=11).value = "Кон. баланс токена"
        worksheet.cell(row=1, column=12).value = "Время"
        workbook.save(log_file)
        workbook.close()

        worksheet = workbook.create_sheet('Market transactions')
        worksheet.cell(row=1, column=1).value = "№ кошелька"
        worksheet.cell(row=1, column=2).value = "№ транзакции"
        worksheet.cell(row=1, column=3).value = "Адрес кошелька"
        worksheet.cell(row=1, column=4).value = "Тип операции"
        worksheet.cell(row=1, column=5).value = "Сумма операции"
        worksheet.cell(row=1, column=6).value = "Hash транзакции"
        worksheet.cell(row=1, column=7).value = "Нач. баланс USDC"
        worksheet.cell(row=1, column=8).value = "Кон. баланс USDC"
        worksheet.cell(row=1, column=9).value = "Время"
        workbook.save(log_file)
        workbook.close()


class LogMarket(object):
    def __init__(self, index, txn_num, address, op_type, operation_value, txn_hash,
                 balance_st_usdc, balance_end_usdc):
        self.address = address
        self.txn_num = txn_num
        self.operation_value = operation_value
        self.op_type = op_type
        self.txn_hash = txn_hash
        self.balance_st_usdc = balance_st_usdc
        self.balance_end_usdc = balance_end_usdc
        self.index = index

    def write_log(self, script_time):
        while True:
            try:
                log_file = settings.log_file
                workbook = openpyxl.load_workbook(log_file)
                worksheet = workbook['Market transactions']
                last_row = worksheet.max_row
                worksheet.cell(row=last_row + 1, column=1).value = self.index
                worksheet.cell(row=last_row + 1, column=2).value = self.txn_num
                worksheet.cell(row=last_row + 1, column=3).value = self.address
                worksheet.cell(row=last_row + 1, column=4).value = self.op_type
                worksheet.cell(row=last_row + 1, column=5).value = self.operation_value
                worksheet.cell(row=last_row + 1, column=6).value = self.txn_hash
                worksheet.cell(row=last_row + 1, column=7).value = self.balance_st_usdc
                worksheet.cell(row=last_row + 1, column=7).number_format = '0.00'
                worksheet.cell(row=last_row + 1, column=8).value = self.balance_end_usdc
                worksheet.cell(row=last_row + 1, column=8).number_format = '0.00'
                worksheet.cell(row=last_row + 1, column=9).value = script_time
                workbook.save(log_file)
                workbook.close()
                break
            except PermissionError:
                cs_logger.info(f'Не получается сохранить файл! Закройте Excel. Нажмите Enter для продолжения...')
                input("")


def get_last_row_overall():
    log_file = settings.log_file
    workbook = openpyxl.load_workbook(log_file)
    worksheet = workbook['Overall']
    last_row = worksheet.max_row
    return last_row


def write_overall(wallet, balance_st, balance_end, script_time, nonce):
    while True:
        try:
            workbook = openpyxl.load_workbook(settings.log_file)
            worksheet = workbook['Overall']
            last_row = settings.last_row
            worksheet.cell(row=last_row + wallet.index, column=1).value = wallet.wallet_num
            worksheet.cell(row=last_row + wallet.index, column=2).value = wallet.address
            worksheet.cell(row=last_row + wallet.index, column=3).value = wallet.exchange_address

            worksheet.cell(row=last_row + wallet.index, column=4).value = wallet.bridge_sum
            worksheet.cell(row=last_row + wallet.index, column=4).number_format = '0.00000'
            worksheet.cell(row=last_row + wallet.index, column=5).value = balance_st
            worksheet.cell(row=last_row + wallet.index, column=5).number_format = '0.00000'
            worksheet.cell(row=last_row + wallet.index, column=6).value = balance_end
            worksheet.cell(row=last_row + wallet.index, column=6).number_format = '0.00000'
            worksheet.cell(row=last_row + wallet.index, column=7).value = wallet.txn_num
            worksheet.cell(row=last_row + wallet.index, column=8).value = nonce

            worksheet.cell(row=last_row + wallet.index, column=9).value = wallet.exc_bal_st
            worksheet.cell(row=last_row + wallet.index, column=9).number_format = '0.00000'

            worksheet.cell(row=last_row + wallet.index, column=10).value = wallet.exc_bal_end
            worksheet.cell(row=last_row + wallet.index, column=10).number_format = '0.00000'

            worksheet.cell(row=last_row + wallet.index, column=11).value = wallet.fwdx_value
            worksheet.cell(row=last_row + wallet.index, column=11).number_format = '0.00'

            worksheet.cell(row=last_row + wallet.index, column=12).value = wallet.zkdx_value
            worksheet.cell(row=last_row + wallet.index, column=12).number_format = '0.00'
            worksheet.cell(row=last_row + wallet.index, column=13).value = script_time
            workbook.save(settings.log_file)
            workbook.close()
            break
        except PermissionError:
            cs_logger.info(f'Не получается сохранить файл! Закройте Excel. Нажмите Enter для продолжения...')
            input("")


def rewrite_overall(wallet, balance_end, nonce):
    while True:
        try:
            workbook = openpyxl.load_workbook(settings.log_file)
            worksheet = workbook['Overall']
            last_row = settings.last_row
            worksheet.cell(row=last_row + wallet.index, column=4).value = wallet.bridge_sum
            worksheet.cell(row=last_row + wallet.index, column=6).value = balance_end
            worksheet.cell(row=last_row + wallet.index, column=7).value = wallet.txn_num
            worksheet.cell(row=last_row + wallet.index, column=8).value = nonce
            worksheet.cell(row=last_row + wallet.index, column=9).value = wallet.exc_bal_st
            worksheet.cell(row=last_row + wallet.index, column=10).value = wallet.exc_bal_end
            worksheet.cell(row=last_row + wallet.index, column=11).value = wallet.fwdx_value
            worksheet.cell(row=last_row + wallet.index, column=12).value = wallet.zkdx_value
            workbook.save(settings.log_file)
            workbook.close()
            break
        except PermissionError:
            cs_logger.info(f'Не получается сохранить файл! Закройте Excel. Нажмите Enter для продолжения...')
            input("")


class LogBridge(object):
    def __init__(self, index, net_from, net_to, address, bridge_value,
                 txn_hash_from, balance_from_st, balance_to_st, balance_from_end, balance_to_end):
        self.address = address
        self.bridge_value = bridge_value
        self.txn_hash_from = txn_hash_from
        self.balance_from_st = balance_from_st
        self.balance_to_st = balance_to_st
        self.balance_from_end = balance_from_end
        self.balance_to_end = balance_to_end
        self.index = index
        self.net_from = net_from
        self.net_to = net_to

    def write_log(self, script_time):
        while True:
            try:
                log_file = settings.log_file
                workbook = openpyxl.load_workbook(log_file)
                worksheet = workbook['Bridge transactions']
                last_row = worksheet.max_row
                worksheet.cell(row=last_row + 1, column=1).value = self.index
                worksheet.cell(row=last_row + 1, column=2).value = f'{self.address}'
                worksheet.cell(row=last_row + 1, column=3).value = f'{self.net_from}'
                worksheet.cell(row=last_row + 1, column=4).value = f'{self.net_to}'
                worksheet.cell(row=last_row + 1, column=5).value = f'{self.txn_hash_from}'
                worksheet.cell(row=last_row + 1, column=6).value = self.balance_from_st
                worksheet.cell(row=last_row + 1, column=6).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=7).value = self.balance_from_end
                worksheet.cell(row=last_row + 1, column=7).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=8).value = self.bridge_value
                worksheet.cell(row=last_row + 1, column=8).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=9).value = self.balance_to_st
                worksheet.cell(row=last_row + 1, column=9).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=10).value = self.balance_to_end
                worksheet.cell(row=last_row + 1, column=10).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=11).value = f'{script_time}'
                workbook.save(log_file)
                workbook.close()
                break
            except PermissionError:
                cs_logger.info(f'Не получается сохранить файл! Закройте Excel. Нажмите Enter для продолжения...')
                input("")

    def rewrite_log(self):
        while True:
            try:
                log_file = settings.log_file
                workbook = openpyxl.load_workbook(log_file)
                worksheet = workbook['Bridge transactions']
                last_row = worksheet.max_row
                worksheet.cell(row=last_row, column=7).value = self.balance_from_end
                worksheet.cell(row=last_row, column=10).value = self.balance_to_end
                workbook.save(log_file)
                workbook.close()
                break
            except PermissionError:
                cs_logger.info(f'Не получается сохранить файл! Закройте Excel. Нажмите Enter для продолжения...')
                input("")


class LogSwap(object):
    def __init__(self, wallet_num, txn_num, address, swapper, swap_value, hash_txn, balance_start_eth,
                 balance_end_eth, balance_start_token, balance_end_token):
        self.wallet_num = wallet_num
        self.txn_num = txn_num
        self.address = address
        self.swapper = swapper
        self.swap_value = swap_value
        self.hash_txn = hash_txn
        self.balance_start_eth = balance_start_eth
        self.balance_end_eth = balance_end_eth
        self.balance_start_token = balance_start_token
        self.balance_end_token = balance_end_token

    def write_log(self, token, script_time):
        while True:
            try:
                log_file = settings.log_file
                workbook = openpyxl.load_workbook(log_file)
                worksheet = workbook['Swap transactions']
                last_row = worksheet.max_row
                worksheet.cell(row=last_row + 1, column=1).value = self.wallet_num
                worksheet.cell(row=last_row + 1, column=2).value = self.txn_num
                worksheet.cell(row=last_row + 1, column=3).value = f'{self.address}'
                worksheet.cell(row=last_row + 1, column=4).value = f'{self.swapper}'

                if token == 1:
                    worksheet.cell(row=last_row + 1, column=5).value = self.swap_value
                    worksheet.cell(row=last_row + 1, column=5).number_format = '0.00000'
                    worksheet.cell(row=last_row + 1, column=6).value = ''
                if token == 2:
                    worksheet.cell(row=last_row + 1, column=5).value = ''
                    worksheet.cell(row=last_row + 1, column=6).value = self.swap_value
                    worksheet.cell(row=last_row + 1, column=6).number_format = '0.00000'

                worksheet.cell(row=last_row + 1, column=7).value = f'{self.hash_txn}'
                worksheet.cell(row=last_row + 1, column=8).value = self.balance_start_eth
                worksheet.cell(row=last_row + 1, column=8).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=9).value = self.balance_end_eth
                worksheet.cell(row=last_row + 1, column=9).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=10).value = self.balance_start_token
                worksheet.cell(row=last_row + 1, column=10).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=11).value = self.balance_end_token
                worksheet.cell(row=last_row + 1, column=11).number_format = '0.00000'
                worksheet.cell(row=last_row + 1, column=12).value = f'{script_time}'
                workbook.save(log_file)
                workbook.close()
                break
            except PermissionError:
                cs_logger.info(f'Не получается сохранить файл! Закройте Excel. Нажмите Enter для продолжения...')
                input("")
